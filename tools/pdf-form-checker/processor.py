"""
PDF Form Field Checker - Browser Version
Uses PyScript to generate Excel reports from comparison results
"""

from pyscript import window
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import base64


def create_excel_report(data):
    """Generate Excel workbook from comparison results"""
    wb = Workbook()
    wb.active.title = "MASTER"
    
    control_name = data['controlName']
    results = data['results']
    
    # Create master summary sheet
    ws = wb.active
    ws["A1"] = "MASTER SUMMARY - PDF Form Field Review"
    ws.merge_cells("A1:E1")
    style_cell(ws["A1"], bg_color="1F4E78", font_color="FFFFFF", bold=True, size=13, center=True)
    ws.row_dimensions[1].height = 22
    
    ws["A3"] = f"Control PDF: {control_name}"
    style_cell(ws["A3"], bold=True, size=10)
    
    headers = ["Review PDF", "Total Fields", "Matching Fields", "Mismatched Fields", "Match %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        style_cell(cell, bg_color="1F4E78", font_color="FFFFFF", bold=True, size=11, center=True)
    
    row = 6
    for result in results:
        total = len(result['matches']) + len(result['mismatches'])
        matches = len(result['matches'])
        
        ws[f"A{row}"] = result['pdfName']
        ws[f"B{row}"] = total
        ws[f"C{row}"] = matches
        ws[f"D{row}"] = len(result['mismatches'])
        
        pct = (matches / total * 100) if total > 0 else 0
        ws[f"E{row}"] = pct / 100
        ws[f"E{row}"].number_format = "0.0%"
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center" if col > 1 else "left"
            )
        
        row += 1
    
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    
    # Create individual sheets for each review PDF
    for result in results:
        create_summary_sheet(wb, result)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def style_cell(cell, bg_color=None, font_color=None, bold=False, size=11, center=False):
    """Apply styling to a cell"""
    if bg_color:
        cell.fill = PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        )
    if font_color:
        cell.font = Font(bold=bold, color=font_color, size=size)
    elif bold or size != 11:
        cell.font = Font(bold=bold, size=size)
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_summary_sheet(wb, result):
    """Create a summary sheet for a review PDF"""
    pdf_name = result['pdfName']
    mismatches = result['mismatches']
    
    # Truncate sheet name to 30 characters
    sheet_name = f"Summary - {pdf_name[:20]}"
    ws = wb.create_sheet(sheet_name)
    
    if not mismatches:
        ws["A1"] = f"Summary - {pdf_name}"
        ws.merge_cells("A1:G1")
        style_cell(ws["A1"], bg_color="00B050", font_color="FFFFFF", bold=True, size=13, center=True)
        ws.row_dimensions[1].height = 22
        
        ws["A3"] = "100% MATCH"
        style_cell(ws["A3"], bg_color="C6EFCE", font_color="006100", bold=True, size=16, center=True)
        ws.row_dimensions[3].height = 30
    else:
        ws["A1"] = f"Summary - {pdf_name}"
        ws.merge_cells("A1:G1")
        style_cell(ws["A1"], bg_color="C00000", font_color="FFFFFF", bold=True, size=13, center=True)
        ws.row_dimensions[1].height = 22
        
        headers = ["Ctrl Pg", "Rev Pg", "Field Name", "Control (Status)", "Review (Status)", "Control Value", "Review Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            style_cell(cell, bg_color="C00000", font_color="FFFFFF", bold=True, size=11, center=True)
        
        row = 4
        for field in mismatches:
            ws[f"A{row}"] = field['cPage']
            ws[f"B{row}"] = field['rPage']
            ws[f"C{row}"] = field['name']
            ws[f"D{row}"] = field['cStatus']
            ws[f"E{row}"] = field['rStatus']
            ws[f"F{row}"] = field['cValue']
            ws[f"G{row}"] = field['rValue']
            
            style_cell(ws[f"D{row}"], bg_color="FFC7CE", center=True)
            style_cell(ws[f"E{row}"], bg_color="FFC7CE", center=True)
            
            row += 1
        
        row += 1
        ws[f"A{row}"] = "SUMMARY"
        style_cell(ws[f"A{row}"], bold=True, size=11)
        row += 1
        ws[f"A{row}"] = "Total Mismatches"
        ws[f"B{row}"] = len(mismatches)
        style_cell(ws[f"A{row}"], bold=True)
    
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 25


async def generate_excel_report(data):
    """Generate and download Excel report"""
    try:
        # Generate workbook
        excel_bytes = create_excel_report(data)
        
        # Convert to base64 for download
        b64 = base64.b64encode(excel_bytes).decode()
        
        # Trigger download
        js_code = f"""
        const link = document.createElement('a');
        link.href = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}';
        link.download = 'PDF_Form_Fields_Report.xlsx';
        document.body.appendChild(link);
        link.click();
        document.body.removeChild(link);
        """
        
        window.eval(js_code)
        
    except Exception as e:
        window.alert(f"Error generating Excel report: {str(e)}")
        raise


# Expose function to JavaScript
window.generateExcelReport = generate_excel_report
